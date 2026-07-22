"""Импорт контактов из CSV / Excel / TXT и валидация адресов.

Импорт устроен в два шага:
  1. read_table(path)      -> (headers, rows)  — сырые строки из файла
  2. import_rows(...)      -> ImportResult      — с маппингом колонок,
                                                  валидацией и дедупликацией.

Так UI может показать превью и дать пользователю сопоставить колонки.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

try:
    from email_validator import EmailNotValidError, validate_email
except Exception:  # библиотека необязательна — есть запасной regex
    validate_email = None
    EmailNotValidError = Exception

# Грубая проверка, если email-validator не установлен.
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def normalize_email(raw: str) -> str | None:
    """Вернуть нормализованный адрес или None, если он невалиден."""
    if not raw:
        return None
    raw = raw.strip().strip("<>").lower()
    if validate_email is not None:
        try:
            # check_deliverability=False — не ходим в DNS при импорте,
            # это медленно и не нужно для базовой очистки.
            info = validate_email(raw, check_deliverability=False)
            return info.normalized.lower()
        except EmailNotValidError:
            return None
    return raw if _EMAIL_RE.match(raw) else None


# ---------------- чтение файла ----------------

def read_table(path: str | Path) -> tuple[list[str], list[list[str]]]:
    """Прочитать таблицу. Возвращает (заголовки, строки).

    Поддержка: .csv, .tsv, .txt (разделители , ; tab), .xlsx.
    Для .txt со списком голых адресов заголовок будет ['email'].
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    return _read_delimited(path)


def _read_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("Для .xlsx нужен пакет openpyxl (pip install openpyxl)") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return [], []
    rows = [[("" if c is None else str(c)) for c in r] for r in rows_iter]
    wb.close()
    return header, rows


def _read_delimited(path: Path) -> tuple[list[str], list[list[str]]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = text[:4096]
    # Определяем разделитель; для простого списка адресов — перевод строки.
    delimiter = None
    for cand in (",", ";", "\t", "|"):
        if cand in sample:
            delimiter = cand
            break
    if delimiter is None:
        # один столбец — считаем, что это адреса
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        # если первая строка похожа на заголовок 'email' — пропустим её как заголовок
        header = ["email"]
        rows = [[ln] for ln in lines]
        if rows and normalize_email(rows[0][0]) is None and rows[0][0].lower() in ("email", "e-mail", "почта"):
            rows = rows[1:]
        return header, rows

    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    all_rows = list(reader)
    if not all_rows:
        return [], []
    header = [h.strip() for h in all_rows[0]]
    rows = [r for r in all_rows[1:] if any(cell.strip() for cell in r)]
    return header, rows


# ---------------- маппинг и импорт ----------------

# Автоопределение колонок по типичным названиям.
_COLUMN_HINTS = {
    "email": ["email", "e-mail", "mail", "почта", "адрес", "e_mail"],
    "first_name": ["first_name", "firstname", "name", "имя", "fname"],
    "last_name": ["last_name", "lastname", "surname", "фамилия", "lname"],
    "company": ["company", "org", "organization", "компания", "организация", "фирма"],
}


def guess_mapping(headers: list[str]) -> dict[str, int]:
    """Сопоставить стандартные поля с индексами колонок по названиям."""
    mapping: dict[str, int] = {}
    lowered = [h.strip().lower() for h in headers]
    for field_name, hints in _COLUMN_HINTS.items():
        for i, h in enumerate(lowered):
            if h in hints:
                mapping[field_name] = i
                break
    # если email не нашли по названию, но колонка одна — берём её
    if "email" not in mapping and len(headers) == 1:
        mapping["email"] = 0
    return mapping


@dataclass
class ImportResult:
    imported: int = 0
    updated: int = 0
    invalid: int = 0
    duplicates: int = 0
    suppressed: int = 0
    invalid_samples: list[str] = field(default_factory=list)

    def summary(self) -> str:
        return (
            f"Добавлено: {self.imported}, обновлено: {self.updated}, "
            f"дублей в файле: {self.duplicates}, невалидных: {self.invalid}, "
            f"в стоп-листе (пропущены): {self.suppressed}"
        )


def import_rows(storage, headers: list[str], rows: list[list[str]],
                mapping: dict[str, int], *, source: str = "") -> ImportResult:
    """Импортировать строки в хранилище согласно маппингу колонок.

    mapping: {'email': idx, 'first_name': idx, ...}. Обязателен ключ 'email'.
    Прочие колонки, не попавшие в стандартные поля, складываются в fields_json.
    Адреса из стоп-листа пропускаются — это ключевое правило белой рассылки.
    """
    if "email" not in mapping:
        raise ValueError("Не указана колонка с email")

    result = ImportResult()
    seen: set[str] = set()
    suppressed = storage.suppressed_set()
    email_idx = mapping["email"]
    std_indices = set(mapping.values())

    def cell(row, idx):
        return row[idx].strip() if idx is not None and idx < len(row) else ""

    for row in rows:
        raw_email = cell(row, email_idx)
        email = normalize_email(raw_email)
        if not email:
            result.invalid += 1
            if len(result.invalid_samples) < 10 and raw_email:
                result.invalid_samples.append(raw_email)
            continue
        if email in seen:
            result.duplicates += 1
            continue
        seen.add(email)
        if email in suppressed:
            result.suppressed += 1
            continue

        # произвольные поля = колонки, не занятые стандартными
        extra = {}
        for i, h in enumerate(headers):
            if i not in std_indices and i < len(row) and row[i].strip():
                key = h.strip() or f"col{i}"
                extra[key] = row[i].strip()

        _id, is_new = storage.upsert_contact(
            email,
            first_name=cell(row, mapping.get("first_name")),
            last_name=cell(row, mapping.get("last_name")),
            company=cell(row, mapping.get("company")),
            fields=extra,
            source=source,
        )
        if is_new:
            result.imported += 1
        else:
            result.updated += 1

    return result
